import os
import pandas as pd
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from datetime import datetime, timedelta
from coinpaprika import client as Coinpaprika
import time
from tabulate import tabulate
from dotenv import load_dotenv
import requests
from vaderSentiment.vaderSentiment import SentimentIntensityAnalyzer
import openai  # Assuming GPT-4 is accessible via OpenAI API
import json
import re
import math
import os
import logging 
from fuzzywuzzy import fuzz
from fuzzywuzzy import process
import san
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import psycopg2
from psycopg2 import Error
import matplotlib.pyplot as plt
import traceback
import pandas as pd
from psycopg2 import sql, OperationalError
import base64
from io import BytesIO
from sqlalchemy import create_engine
from sqlalchemy.exc import SQLAlchemyError
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from email.mime.image import MIMEImage  # This will fix the NameError

# Initialize Sanpy API key
SAN_API_KEY = os.getenv('SAN_API_KEY')
san.ApiConfig.api_key = SAN_API_KEY

surge_words = [
    "surge", "spike", "soar", "rocket", "skyrocket", "rally", "boom", "bullish", 
    "explosion", "rise", "uptrend", "bull run", "moon", "parabolic", "spurt", 
    "climb", "jump", "upswing", "gain", "increase", "growth", "rebound", 
    "breakout", "spurt", "pump", "fly", "explode", "shoot up", "hike", 
    "expand", "appreciate", "bull market", "peak", "momentum", "outperform", 
    "spike up", "ascend", "elevation", "expansion", "revive", "uprising", 
    "push up", "escalate", "rise sharply", "escalation", "recover", 
    "inflation", "strengthen", "gain strength", "intensify"
]

# Volume thresholds for liquidity risk
LOW_VOLUME_THRESHOLD_LARGE = 1_000_000  # Large-cap coins with daily volume under $1M
LOW_VOLUME_THRESHOLD_MID = 500_000  # Mid-cap coins with daily volume under $500k
LOW_VOLUME_THRESHOLD_SMALL = 100_000  # Small-cap coins with daily volume under $100k

# Configure logging
logging.basicConfig(
    filename='monitor_coins.log',  # Path to the log file
    level=logging.DEBUG,  # Set the logging level (DEBUG, INFO, WARNING, ERROR, CRITICAL)
    format='%(asctime)s - %(levelname)s - %(message)s'
)
# Load environment variables from .env file
load_dotenv()

# Initialize the CoinPaprika client
client = Coinpaprika.Client(api_key=os.getenv('COIN_PAPRIKA_API_KEY'))
openai.api_key = os.getenv('OPENAI_API_KEY')

# Initialize the VADER sentiment analyzer
analyzer = SentimentIntensityAnalyzer()

# Configuration
EMAIL_FROM = os.getenv('EMAIL_FROM') 
EMAIL_TO = os.getenv('EMAIL_TO') 
SMTP_SERVER = os.getenv('SMTP_SERVER')  
SMTP_USERNAME = os.getenv('SMTP_USERNAME')
SMTP_PASSWORD = os.getenv('SMTP_PASSWORD')
SMTP_PORT = 587
SMTP_USERNAME = os.getenv('SMTP_USERNAME')
SMTP_PASSWORD = os.getenv('SMTP_PASSWORD')
RESULTS_FILE = "surging_coins.csv"
CRYPTO_NEWS_TICKERS = "tickers.csv"
FEAR_GREED_THRESHOLD = 60  # Fear and Greed index threshold
HIGH_VOLATILITY_THRESHOLD = 0.05  # 5% volatility is considered high
MEDIUM_VOLATILITY_THRESHOLD = 0.02  # 2% volatility is considered medium

# Flag to test with just Bitcoin
TEST_ONLY = False  # Set to False to monitor all coins
MAX_RETRIES = 2  # Maximum number of retries for API calls
BACKOFF_FACTOR = 2  # Factor by which the wait time increases after each failure

CUMULATIVE_SCORE_REPORTING_THRESHOLD=50 # Only report results with cumulative score above this % value

def api_call_with_retries(api_function, *args, **kwargs):
    """
    Calls an API function with retries in case of failure.

    Args:
        api_function (function): The API function to call.
        *args: Arguments to pass to the API function.
        **kwargs: Keyword arguments to pass to the API function.

    Returns:
        The result of the API function call.

    Raises:
        Exception: If the API function call fails after MAX_RETRIES retries.
    """
    retries = 0
    wait_time = 60  # Initial wait time (seconds)
    time.sleep(5)

    while retries < MAX_RETRIES:
        try:
            return api_function(*args, **kwargs)
        except Exception as e:
            retries += 1
            if retries < MAX_RETRIES:
                logging.debug(f"Error during API call: {e}. Retrying in {wait_time} seconds...")
                time.sleep(wait_time)
                wait_time *= BACKOFF_FACTOR
            else:
                logging.debug(f"Max retries reached. Failed to complete API call.")
                raise

def fetch_fear_and_greed_index():
    """
    Fetches the current Fear and Greed Index value from the Alternative.me API.

    Returns:
        int: The current Fear and Greed Index value, or None if the API call fails.
    """
    try:
        response = api_call_with_retries(requests.get, 'https://api.alternative.me/fng/')
        data = response.json()
        return int(data['data'][0]['value'])
    except Exception as e:
        logging.debug(f"Error fetching Fear and Greed Index: {e}")
        return None

def fetch_historical_ticker_data(coin_id, start_date, end_date):
    """
    Fetches historical ticker data for the specified coin and date range.

    Parameters:
        coin_id (str): The CoinPaprika ID of the cryptocurrency.
        start_date (str): The start date of the period for which to fetch the ticker data (in YYYY-MM-DD format).
        end_date (str): The end date of the period for which to fetch the ticker data (in YYYY-MM-DD format).

    Returns:
        pandas.DataFrame: A DataFrame containing the historical ticker data, with columns 'date', 'price', 'coin_id', 'volume_24h', and 'market_cap'. 
                          If the API call fails, an empty DataFrame is returned.
    """
    try:
        historical_ticker = api_call_with_retries(client.historical, coin_id=coin_id, start=start_date, end=end_date, interval="1d", quote="usd")

        if isinstance(historical_ticker, list) and historical_ticker:
            df = pd.DataFrame(historical_ticker)

            if 'price' in df.columns and 'volume_24h' in df.columns:
                df['date'] = pd.to_datetime(df['timestamp']).dt.date
                df['coin_id'] = coin_id
                df = df[['date', 'price', 'coin_id', 'volume_24h', 'market_cap']]
                df = df.sort_values(by='date')  # Sort the DataFrame by the 'date' column
                return df
            else:
                logging.debug(f"Missing expected columns in historical data for {coin_id}.")
                return pd.DataFrame()
        else:
            logging.debug(f"Unexpected data format returned for {coin_id}.")
            return pd.DataFrame()

    except Exception as e:
        logging.debug(f"Error fetching historical data for {coin_id}: {e}")
        return pd.DataFrame()

        
def filter_active_and_ranked_coins(coins, max_coins, rank_threshold=1000):
    """
    Filters the list of coins by rank, activity status, and new status, selecting up to max_coins.

    Parameters:
        coins (list): List of coins with rank, activity status, and new status information.
        max_coins (int): The maximum number of coins to return.
        rank_threshold (int): The maximum rank a coin must have to be included (e.g., rank <= 1000).

    Returns:
        list: Filtered list of coins that are active, not new, and ranked within the rank_threshold.
    """
    # Filter out coins that are not active, are new, or have a rank above the rank_threshold
    active_ranked_coins = [coin for coin in coins if coin.get('is_active', False) and not coin.get('is_new', True) and coin.get('rank', None) <= rank_threshold]

    # Limit the list to max_coins
    return active_ranked_coins[:max_coins]


def fetch_coin_events(coin_id):
    """
    Fetches events for a given cryptocurrency from the CoinPaprika API.

    Parameters:
        coin_id (str): The ID of the cryptocurrency.

    Returns:
        list: A list of recent events for the cryptocurrency, or an empty list if the API call fails.
    """
    try:
        events = api_call_with_retries(client.events, coin_id=coin_id)
        
        if not events:
            logging.debug(f"No events found for {coin_id}.")
            return []

        # Filter events to include only those from the past week and exclude future dates
        one_week_ago = datetime.now() - timedelta(days=7)
        recent_events = []

        for event in events:
            event_date = datetime.strptime(event['date'], "%Y-%m-%dT%H:%M:%SZ")
            if one_week_ago <= event_date <= datetime.now():
                recent_events.append(event)

        logging.debug(f"Events found for {coin_id}: {len(recent_events)} recent events")
        return recent_events

    except Exception as e:
        logging.debug(f"Error fetching events for {coin_id}: {e}")
        return []

def fetch_twitter_data(coin_id):
    """
    Fetches tweets for a given cryptocurrency from the CoinPaprika API.

    Parameters:
        coin_id (str): The ID of the cryptocurrency.

    Returns:
        pd.DataFrame: A pandas DataFrame containing tweets for the cryptocurrency from the past week, or an empty DataFrame if the API call fails or no tweets are found.
    """
    tweets = api_call_with_retries(client.twitter, coin_id)
    
    if not tweets:
        logging.debug(f"No tweets found for {coin_id}.")
        return pd.DataFrame()

    df = pd.DataFrame(tweets)

    if 'status' not in df.columns or 'date' not in df.columns:
        logging.debug(f"'status' or 'date' column not found in Twitter data for {coin_id}. Available columns: {df.columns.tolist()}")
        return pd.DataFrame()

    # Filter tweets from the past week
    one_week_ago = datetime.now() - timedelta(days=7)  # Naive datetime
    df['date'] = pd.to_datetime(df['date']).dt.tz_localize(None)  # Convert to naive datetime
    df = df[df['date'] >= one_week_ago]

    if df.empty:
        logging.debug(f"No recent tweets (past week) found for {coin_id}.")
        return pd.DataFrame()

    logging.debug(f"Tweets found for {coin_id} in the past week: {len(df)} tweets")

    return df

def load_existing_results():
    """
    Loads existing results from a CSV file.

    Returns:
        pd.DataFrame: Existing results read from the CSV file, or an empty DataFrame if the file does not exist.
    """
    if os.path.exists(RESULTS_FILE):
        return pd.read_csv(RESULTS_FILE)
    else:
        return pd.DataFrame()

def save_result_to_csv(result):
    """
    Saves a single result as a row in a CSV file.

    The result will be appended to the existing file if it exists, or written to a new file if not.

    Parameters:
        result (dict): A dictionary containing at least the keys 'coin', 'market_cap', 'volume_24h', 'price_change_7d', and 'fear_greed_index'.
    """
    if not os.path.exists(RESULTS_FILE):
        pd.DataFrame([result]).to_csv(RESULTS_FILE, mode='w', header=True, index=False)
    else:
        pd.DataFrame([result]).to_csv(RESULTS_FILE, mode='a', header=False, index=False)

def classify_market_cap(market_cap):
    """
    Classify a market capitalization as "Large", "Mid", or "Small".

    Parameters:
        market_cap (int): The market capitalization of the cryptocurrency.

    Returns:
        str: The classification of the market capitalization.
    """
    if market_cap > 10_000_000_000:
        return "Large"
    elif market_cap > 1_000_000_000:
        return "Mid"
    else:
        return "Small"

def classify_volatility(volatility):
    """
    Classify a volatility value as "High", "Medium", or "Low".

    Parameters:
        volatility (float): The volatility of the cryptocurrency.

    Returns:
        str: The classification of the volatility.
    """
    if volatility > HIGH_VOLATILITY_THRESHOLD:
        return "High"
    elif volatility > MEDIUM_VOLATILITY_THRESHOLD:
        return "Medium"
    else:
        return "Low"


def save_cumulative_score_to_aurora(coin_id, coin_name, cumulative_score):
    """
    Save a cumulative score for a specific coin in Amazon Aurora (PostgreSQL) with a date-based timestamp.

    Parameters:
        coin_id (str): The unique identifier for the coin.
        coin_name (str): The name of the coin.
        cumulative_score (float): The cumulative score of the coin.
    """
    connection = None  # Initialize connection variable
    cursor = None  # Initialize cursor variable
    
    try:
        # Establish connection to PostgreSQL Aurora instance
        connection = psycopg2.connect(
            host=os.getenv('AURORA_HOST'),
            database=os.getenv('AURORA_DB'),
            user=os.getenv('AURORA_USER'),
            password=os.getenv('AURORA_PASSWORD'),
            port=os.getenv('AURORA_PORT', 5432)  # Default port for PostgreSQL is 5432
        )
        
        cursor = connection.cursor()

        # Insert the cumulative score with the current date (no time part)
        insert_query = """
            INSERT INTO coin_data (coin_id, coin_name, cumulative_score, timestamp)
            VALUES (%s, %s, %s, %s)
            ON CONFLICT (coin_id, timestamp) 
            DO UPDATE SET cumulative_score = EXCLUDED.cumulative_score;
        """
        
        # Truncate timestamp to just the day (remove time component)
        current_date = datetime.now().date()  # Get only the date part
        
        cursor.execute(insert_query, (coin_id, coin_name, cumulative_score, current_date))
        
        connection.commit()
        print(f"Cumulative score for {coin_name} saved/updated successfully for {current_date}.")
    
    except psycopg2.OperationalError as e:
        print(f"Error connecting to Amazon Aurora DB: {e}")
    
    finally:
        # Check if cursor was created and close it
        if cursor is not None:
            try:
                cursor.close()
                print("Cursor is closed.")
            except Exception as e:
                print(f"Error closing cursor: {e}")

        # Check if connection was created and close it
        if connection is not None:
            try:
                connection.close()
                print("PostgreSQL connection is closed.")
            except Exception as e:
                print(f"Error closing connection: {e}")


def get_volume_thresholds(market_cap_class, volatility_class):
    """
    Returns the volume thresholds for a given market capitalization class and volatility class.

    Parameters:
        market_cap_class (str): The market capitalization class, one of "Large", "Mid", or "Small".
        volatility_class (str): The volatility class, one of "High", "Medium", or "Low".

    Returns:
        tuple: A tuple of six floats, representing the volume change thresholds for short-term, medium-term, and long-term periods, respectively.
    """
    thresholds = {
        ("Large", "High"): (2, 4, 1.5, 3, 1.2, 2),
        ("Large", "Medium"): (1.5, 3, 1.2, 2, 1.1, 1.5),
        ("Large", "Low"): (1.2, 2, 1.1, 1.5, 1, 1.2),
        ("Mid", "High"): (3, 6, 2, 4, 1.5, 2.5),
        ("Mid", "Medium"): (2, 4, 1.5, 3, 1.2, 2),
        ("Mid", "Low"): (1.5, 3, 1.2, 2, 1, 1.5),
        ("Small", "High"): (5, 10, 3, 6, 2, 4),
        ("Small", "Medium"): (3, 6, 2, 4, 1.5, 2.5),
        ("Small", "Low"): (2, 4, 1.5, 3, 1.2, 2)
    }
    return thresholds.get((market_cap_class, volatility_class), (2, 4, 1.5, 3, 1.2, 2))

def analyze_volume_change(volume_data, market_cap, volatility):
    """
    Analyze the volume changes of a cryptocurrency over three time periods.

    Parameters:
        volume_data (pd.DataFrame): Volume data for the cryptocurrency.
        market_cap (int): Market capitalization of the cryptocurrency.
        volatility (float): Volatility of the cryptocurrency.

    Returns:
        tuple: The volume change score and an explanation string detailing
               which periods had significant volume changes.
    """
    # Classify the market capitalization and volatility
    market_cap_class = classify_market_cap(market_cap)
    volatility_class = classify_volatility(volatility)

    # Get volume thresholds based on the market cap and volatility classification
    short_term_threshold, short_term_max, medium_term_threshold, medium_term_max, long_term_threshold, long_term_max = get_volume_thresholds(market_cap_class, volatility_class)

    # Analyze short-term, medium-term, and long-term volume changes
    short_term_change = calculate_volume_change(volume_data, period="short")
    medium_term_change = calculate_volume_change(volume_data, period="medium")
    long_term_change = calculate_volume_change(volume_data, period="long")

    volume_score = 0
    explanation_parts = []

    if short_term_change > short_term_threshold and short_term_change < short_term_max:
        volume_score += 1
        explanation_parts.append(f"Short-term volume change of {short_term_change*100:.2f}% exceeded the threshold of {short_term_threshold*100:.2f}%")
    
    if medium_term_change > medium_term_threshold and medium_term_change < medium_term_max:
        volume_score += 1
        explanation_parts.append(f"Medium-term volume change of {medium_term_change*100:.2f}% exceeded the threshold of {medium_term_threshold*100:.2f}%")
    
    if long_term_change > long_term_threshold and long_term_change < long_term_max:
        volume_score += 1
        explanation_parts.append(f"Long-term volume change of {long_term_change*100:.2f}% exceeded the threshold of {long_term_threshold*100:.2f}%")

    # Combine the explanation parts into a single string
    explanation = " | ".join(explanation_parts) if explanation_parts else "No significant volume changes detected."

    return volume_score, explanation


def has_consistent_weekly_growth(historical_df):
    # Calculate daily price changes
    """
    Returns True if the given historical data shows consistent weekly growth,
    which is defined as at least 4 out of the last 7 days having a positive
    price change.

    Parameters:
        historical_df (pd.DataFrame): A pandas DataFrame containing historical
            data for the cryptocurrency, with columns for 'date', 'price', and
            'volume_24h'.

    Returns:
        bool: True if the price has been growing for at least 4 of the last 7
            days, False otherwise.
    """
    historical_df['price_change'] = historical_df['price'].pct_change()

    # Filter the last 7 days
    last_week_df = historical_df.tail(7)
    
    # Count how many days had a positive price change
    rising_days = last_week_df[last_week_df['price_change'] > 0].shape[0]
    
    # Consider consistent growth if at least 4 out of 7 days were rising
    return rising_days >= 4

def fetch_santiment_slugs():
    """
    Fetch the available slugs from Santiment using the sanpy API.

    Returns:
        pd.DataFrame: DataFrame containing Santiment slugs and project information.
    """
    try:
        # Fetch available slugs using sanpy API
        all_projects = san.get(
            "projects/all",
            interval="1d",
            columns=["slug", "name", "ticker", "infrastructure", "mainContractAddress"]
        )
        projects_df = pd.DataFrame(all_projects)

        # Normalize the coin names for matching
        projects_df['name_normalized'] = projects_df['name'].apply(lambda x: re.sub(r'\W+', '', x.lower()))

        logging.info(f"Fetched {len(projects_df)} Santiment slugs")
        return projects_df

    except Exception as e:
        logging.error(f"Error fetching Santiment slugs: {e}")
        return pd.DataFrame()  # Return empty DataFrame on error



def match_coins_with_santiment(coin_name, santiment_slugs_df):
    """
    Matches a given coin name with the Santiment slugs dataframe.
    
    Parameters:
    coin_name (str): The name of the coin to match.
    santiment_slugs_df (pd.DataFrame): The dataframe containing Santiment slugs and normalized names.
    
    Returns:
    str: The Santiment slug if a match is found, else None.
    """    
    # Check if the 'name_normalized' column exists in the dataframe
    if 'name_normalized' not in santiment_slugs_df.columns:
        logging.warning(f"'name_normalized' column not found in the Santiment slugs dataframe.")
        return None

    # Look for exact matches in the normalized names
    match = santiment_slugs_df[santiment_slugs_df['name_normalized'] == coin_name]
    
    if not match.empty:
        return match['slug'].values[0]  # Return the first matching slug
    else:
        logging.info(f"No match found for {coin_name} in Santiment slugs.")
    
    return None


def has_sustained_volume_growth(historical_df):
    # Calculate daily volume changes
    """
    Returns True if the given historical data shows sustained volume growth,
    which is defined as at least 4 out of the last 7 days having a positive
    volume change.

    Parameters:
        historical_df (pd.DataFrame): A pandas DataFrame containing historical
            data for the cryptocurrency, with columns for 'date', 'price', and
            'volume_24h'.

    Returns:
        bool: True if the volume has been growing for at least 4 of the last 7
            days, False otherwise.
    """
    historical_df['volume_change'] = historical_df['volume_24h'].pct_change()

    # Filter the last 7 days
    last_week_df = historical_df.tail(7)
    
    # Count how many days had a positive volume change
    rising_volume_days = last_week_df[last_week_df['volume_change'] > 0].shape[0]
    
    # Consider sustained growth if at least 4 out of 7 days had rising volume
    return rising_volume_days >= 4


def compute_sentiment_for_coin(coin_name, news_data):
    """
    Computes the sentiment score for a given coin based on its news data.

    Parameters:
        coin_name (str): The name of the coin.
        news_data (list): A list of news items related to the coin, where each item is a dict with 'title' and 'description' keys.

    Returns:
        int: 1 if the average sentiment score is very positive (e.g., greater than 0.5), otherwise 0.
    """
    sentiments = []
    for news_item in news_data:
        description = news_item.get('description', '')
        
        # Ensure description is a string and is not empty
        if isinstance(description, str) and description.strip():
            sentiment_score = analyzer.polarity_scores(description)['compound']
            sentiments.append(sentiment_score)
    
    if sentiments:
        average_sentiment = sum(sentiments) / len(sentiments)
    else:
        average_sentiment = 0

    # Return 1 if the average sentiment is very positive (e.g., greater than 0.5), otherwise return 0
    return 1 if average_sentiment > 0.5 else 0

def score_surge_words(news_df, surge_words):
    """
    Analyze news articles and score the presence of surge words.

    Parameters:
        news_df (pd.DataFrame): A pandas DataFrame containing news articles.
        surge_words (list): A list of words indicating a surge in the market.

    Returns:
        tuple: A tuple containing the average surge score across all news articles 
               and a detailed explanation of which articles contributed to the score.
    """
    total_surge_score = 0
    news_count = 0
    explanation = []

    if not news_df.empty:
        for _, news_item in news_df.iterrows():
            description = news_item.get('description', '')

            # Ensure description is a string and is not None
            if isinstance(description, str) and description.strip():
                surge_score = 0
                article_explanation = []

                for word in surge_words:
                    # Fuzzy match each surge word with the description
                    match_score = fuzz.partial_ratio(word.lower(), description.lower())

                    # Add to surge_score based on the match score (scale 0 to 100)
                    if match_score > 75:  # Threshold for a significant match
                        surge_score += match_score / 100.0  # Normalize the score to 0-1 range
                        article_explanation.append(f"Matched word '{word}' with score {match_score}%")

                if surge_score > 0:
                    explanation.append(f"Article: '{news_item.get('title', '')}' contributed to surge score with details: {', '.join(article_explanation)}")

                total_surge_score += surge_score
                news_count += 1

    if news_count > 0:
        average_surge_score = total_surge_score / news_count
    else:
        average_surge_score = 0.0

    return int(math.ceil(average_surge_score)), explanation

def get_fuzzy_trending_score(coin_id, coin_name, trending_coins_scores):
    """
    Analyze trending coins scores and return the maximum score if a fuzzy match is found 
    between the coin ID or name and any of the trending coin tickers.

    Parameters:
        coin_id (str): The CoinPaprika ID of the cryptocurrency.
        coin_name (str): The full name of the cryptocurrency.
        trending_coins_scores (dict): A dictionary with tickers as keys and their respective scores.

    Returns:
        int: The maximum score if a fuzzy match is found, otherwise 0.
    """
    max_score = 0
    for ticker, score in trending_coins_scores.items():
        match_id = fuzz.partial_ratio(ticker.lower(), coin_id.lower())
        match_name = fuzz.partial_ratio(ticker.lower(), coin_name.lower())
        if match_id > 80 or match_name > 80:  # Threshold for considering a match
            max_score = max(max_score, score)
    return max_score


def classify_liquidity_risk(volume_24h, market_cap_class):
    """
    Classify liquidity risk based on trading volume.

    Parameters:
        volume_24h (float): The 24-hour trading volume of the cryptocurrency.
        market_cap_class (str): The market capitalization class, one of "Large", "Mid", or "Small".

    Returns:
        str: The classification of liquidity risk ('Low', 'Medium', 'High').
    """
    if market_cap_class == "Large":
        if volume_24h < LOW_VOLUME_THRESHOLD_LARGE:
            return "High"
        elif volume_24h < LOW_VOLUME_THRESHOLD_LARGE * 2:
            return "Medium"
        else:
            return "Low"
    elif market_cap_class == "Mid":
        if volume_24h < LOW_VOLUME_THRESHOLD_MID:
            return "High"
        elif volume_24h < LOW_VOLUME_THRESHOLD_MID * 2:
            return "Medium"
        else:
            return "Low"
    else:  # Small market cap
        if volume_24h < LOW_VOLUME_THRESHOLD_SMALL:
            return "High"
        elif volume_24h < LOW_VOLUME_THRESHOLD_SMALL * 2:
            return "Medium"
        else:
            return "Low"

def compute_santiment_score_with_thresholds(santiment_data):
    """
    Computes a binary score using Santiment data by applying thresholds for each metric and provides explanations for each score.

    Parameters:
        santiment_data (dict): A dictionary with Santiment metrics data.

    Returns:
        tuple: A final score based on whether each metric exceeds its threshold and an explanation detailing the scoring.
    """
    # Define thresholds for each metric (binary scoring: 0 or 1)
    thresholds = {
        'dev_activity': 10,             # Development activity increase must be greater than 10%
        'daily_active_addresses': 5,    # Daily active addresses increase must be greater than 5%
    }

    # Extract metric values, defaulting to 0 if not available
    dev_activity = santiment_data.get('dev_activity_increase', 0)
    daily_active_addresses = santiment_data.get('daily_active_addresses_increase', 0)

    # Apply thresholds to compute binary scores (0 or 1) and explanations
    explanations = []
 
    if dev_activity > thresholds['dev_activity']:
        dev_activity_score = 1
        explanations.append(f"Development activity increase is significant: {dev_activity}% (Threshold: {thresholds['dev_activity']}%)")
    else:
        dev_activity_score = 0
        explanations.append(f"Development activity increase is low: {dev_activity}% (Threshold: {thresholds['dev_activity']}%)")

    if daily_active_addresses > thresholds['daily_active_addresses']:
        daily_active_addresses_score = 1
        explanations.append(f"Daily active addresses show growth: {daily_active_addresses}% (Threshold: {thresholds['daily_active_addresses']}%)")
    else:
        daily_active_addresses_score = 0
        explanations.append(f"Daily active addresses growth is weak: {daily_active_addresses}% (Threshold: {thresholds['daily_active_addresses']}%)")
    
    # Sum up the scores to get a total score
    total_santiment_score = (
        dev_activity_score +
        daily_active_addresses_score )

    explanation = " | ".join(explanations)  # Combine explanations into a single string

    return total_santiment_score, explanation

def analyze_coin(coin_id, coin_name, end_date, news_df, digest_tickers, trending_coins_scores, santiment_slugs_df):
    """
    Analyzes a given cryptocurrency and returns a dictionary with various analysis scores, 
    including a score for whether the coin appears in the Sundown Digest and trending coins list.

    Parameters:
        coin_id (str): The CoinPaprika ID of the cryptocurrency.
        coin_name (str): The full name of the cryptocurrency.
        end_date (str): The end date of the period for which to fetch the historical ticker data (in YYYY-MM-DD format).
        news_df (pd.DataFrame): A DataFrame containing news articles related to the cryptocurrency.
        digest_tickers (list): A list of tickers extracted from the Sundown Digest.
        trending_coins_scores (dict): A dictionary with tickers as keys and their respective scores.
        santiment_slugs_df (pd.DataFrame): DataFrame containing Santiment slugs for various coins.

    Returns:
        dict: A dictionary with the analysis scores, cumulative score, market cap, volume, and detailed explanation.
    """
    short_term_window = 7
    medium_term_window = 30
    long_term_window = 90
    
    start_date_short_term = (datetime.now() - timedelta(days=short_term_window)).strftime('%Y-%m-%d')
    start_date_medium_term = (datetime.now() - timedelta(days=medium_term_window)).strftime('%Y-%m-%d')
    start_date_long_term = (datetime.now() - timedelta(days=long_term_window)).strftime('%Y-%m-%d')

    historical_df_short_term = fetch_historical_ticker_data(coin_id, start_date_short_term, end_date)
    historical_df_medium_term = fetch_historical_ticker_data(coin_id, start_date_medium_term, end_date)
    historical_df_long_term = fetch_historical_ticker_data(coin_id, start_date_long_term, end_date)

    # Match the coin with Santiment slugs
    santiment_slug = match_coins_with_santiment(coin_name, santiment_slugs_df)

    # If a Santiment slug is found, fetch Santiment data for the last 30 days
    if santiment_slug:
        santiment_data = fetch_santiment_data_for_coin(santiment_slug)
    else:
        santiment_data = {"dev_activity_increase": 0, "daily_active_addresses_increase": 0}


    if 'price' not in historical_df_long_term.columns or historical_df_long_term.empty:
        logging.debug(f"No valid price data available for {coin_id}.")
        return {"coin_id": coin_id, "coin_name": coin_name, "explanation": f"No valid price data available for {coin_id}."}

    twitter_df = fetch_twitter_data(coin_id)
    tweet_score = 1 if not twitter_df.empty else 0

    volatility = historical_df_long_term['price'].pct_change().std()

    # Get market cap and volume from the most recent entry)
    most_recent_market_cap = int(historical_df_long_term['market_cap'].iloc[-1])
    most_recent_volume_24h = int(historical_df_long_term['volume_24h'].iloc[-1])

    # Get price change score and detailed explanation
    price_change_score, price_change_explanation = analyze_price_change(historical_df_long_term['price'], most_recent_market_cap, volatility)

    volume_score, volume_explanation = analyze_volume_change(historical_df_long_term['volume_24h'], most_recent_market_cap, volatility)

    consistent_growth = has_consistent_weekly_growth(historical_df_short_term)
    consistent_growth_score = 1 if consistent_growth else 0

    sustained_volume_growth = has_sustained_volume_growth(historical_df_short_term)
    sustained_volume_growth_score = 1 if sustained_volume_growth else 0

    fear_and_greed_index = fetch_fear_and_greed_index()
    fear_and_greed_score = 1 if fear_and_greed_index is not None and fear_and_greed_index > FEAR_GREED_THRESHOLD else 0

    events = fetch_coin_events(coin_id)
    recent_events_count = sum(1 for event in events if datetime.strptime(event['date'], '%Y-%m-%d') <= datetime.now())
    event_score = 1 if recent_events_count > 0 else 0

    # Classify market cap
    market_cap_class = classify_market_cap(most_recent_market_cap)

    # Calculate liquidity risk based on the most recent 24-hour volume
    liquidity_risk = classify_liquidity_risk(most_recent_volume_24h, market_cap_class)

    # Integrate sentiment analysis
    if not news_df.empty:
        coin_news = news_df[news_df['coin'] == coin_name]
        sentiment_score = compute_sentiment_for_coin(coin_name, coin_news.to_dict('records'))

        # Integrate surge word analysis
        surge_score, surge_explanation = score_surge_words(coin_news, surge_words)
    else:
        sentiment_score = 0
        surge_score = 0
        surge_explanation = "No significant surge-related news detected."

    # Check if the coin is in the digest tickers
    digest_score = 1 if any(fuzz.partial_ratio(ticker.lower(), coin_id.lower()) > 80 or fuzz.partial_ratio(ticker.lower(), coin_name.lower()) > 80 for ticker in digest_tickers) else 0

    # Check if the coin is trending
    trending_score = get_fuzzy_trending_score(coin_id, coin_name, trending_coins_scores)

    # Incorporate Santiment data into the cumulative score
    santiment_score, santiment_explanation = compute_santiment_score_with_thresholds(santiment_data)

    # Calculate cumulative score and its percentage of the maximum score
    cumulative_score = (
        volume_score + tweet_score + consistent_growth_score + sustained_volume_growth_score + 
        fear_and_greed_score + event_score + price_change_score + sentiment_score + surge_score +
        digest_score + trending_score + santiment_score
    )

    # Maximum possible score (adjust as necessary)
    max_possible_score = 13

    # Calculate the cumulative score as a percentage
    cumulative_score_percentage = (cumulative_score / max_possible_score) * 100

    # Build the explanation string, including Santiment data and price change explanation
    explanation = f"{coin_name} ({coin_id}) analysis: "
    explanation += f"Liquidity Risk: {liquidity_risk}, "
    explanation += f"Price Change Score: {'Significant' if price_change_score else 'No significant change'} ({price_change_explanation}), "
    explanation += f"Volume Change Score: {'Significant' if volume_score else 'No significant change'} ({volume_explanation}), "
    explanation += f"Tweets: {'Yes' if tweet_score else 'None'}, "
    explanation += f"Consistent Price Growth: {'Yes' if consistent_growth_score else 'No'}, "
    explanation += f"Sustained Volume Growth: {'Yes' if sustained_volume_growth_score else 'No'}, "
    explanation += f"Fear and Greed Index: {fear_and_greed_index if fear_and_greed_index is not None else 'N/A'}, "
    explanation += f"Recent Events: {recent_events_count}, "
    explanation += f"Sentiment Score: {sentiment_score}, "
    explanation += f"Surge Keywords Score: {surge_score} ({surge_explanation}), "
    explanation += f"Santiment Score: {santiment_score} ({santiment_explanation}), "
    explanation += f"News Digest Score: {digest_score}, "
    explanation += f"Trending Score: {trending_score}, "
    explanation += f"Market Cap: {most_recent_market_cap}, "
    explanation += f"Volume (24h): {most_recent_volume_24h}, "
    explanation += f"Cumulative Surge Score: {cumulative_score} ({cumulative_score_percentage:.2f}%)"

    return {
        "coin_id": coin_id,
        "coin_name": coin_name,
        "market_cap": most_recent_market_cap,  # Add market cap to output
        "volume_24h": most_recent_volume_24h,  # Add volume (24h) to output
        "price_change_score": f"{price_change_score}",
        "volume_change_score": f"{volume_score}",
        "tweets": len(twitter_df) if tweet_score else "None",
        "consistent_growth": "Yes" if consistent_growth_score else "No",
        "sustained_volume_growth": "Yes" if sustained_volume_growth_score else "No",
        "fear_and_greed_index": fear_and_greed_index if fear_and_greed_index is not None else "N/A",
        "events": recent_events_count,
        "sentiment_score": sentiment_score,
        "surging_keywords_score": surge_score,
        "news_digest_score": digest_score,
        "trending_score": trending_score,
        "liquidity_risk": liquidity_risk,
        "santiment_score": santiment_score,
        "cumulative_score": cumulative_score,
        "cumulative_score_percentage": round(cumulative_score_percentage, 2),  # Rounded to 2 decimal places
        "explanation": explanation
    }




def load_tickers(file_path):
    """
    Loads a CSV file containing coin names and tickers, and returns a dictionary mapping
    the coin names to their tickers.

    Parameters:
        file_path (str): The path to the CSV file to load.

    Returns:
        dict: A dictionary mapping coin names to their tickers.
    """
    tickers_df = pd.read_csv(file_path)
    # Create a dictionary mapping the coin names to their tickers
    tickers_dict = pd.Series(tickers_df['Ticker'].values, index=tickers_df['Name']).to_dict()
    return tickers_dict

# Fetch news using the tickers from the loaded CSV
def fetch_news_for_past_week(tickers_dict):
    """
    Fetches news for each coin in the given tickers dictionary for the past week.

    Parameters:
        tickers_dict (dict): A dictionary mapping the coin names to their tickers.

    Returns:
        pd.DataFrame: A pandas DataFrame containing the news articles fetched from the API, with columns 'coin', 'date', 'title', 'description', 'url', and 'source'.
    """
    all_news = []

    end_date = datetime.now().date()
    start_date = end_date - timedelta(days=7)

    for coin_name, coin_ticker in tickers_dict.items():
        logging.debug(f"Fetching news for {coin_name} ({coin_ticker})...")

        formatted_date = end_date.strftime('%Y%m%d')
        week_start = (end_date - timedelta(days=3)).strftime('%m%d%Y')
        week_end = end_date.strftime('%m%d%Y')
        date_str = f"{week_start}-{week_end}"

        url = f"https://cryptonews-api.com/api/v1?tickers={coin_ticker}&items=1&date={date_str}&sortby=rank&token={os.getenv('CRYPTO_NEWS_API_KEY')}"
        response = requests.get(url)
        if response.status_code == 200:
                news_data = response.json()
                if 'data' in news_data and news_data['data']:
                    for article in news_data['data']:
                        all_news.append({
                            "coin": coin_name,
                            "date": formatted_date,  # Log the specific day
                            "title": article["title"],
                            "description": article.get("text", ""),
                            "url": article["news_url"],
                            "source": article["source_name"]
                        })
                else:
                    logging.debug(f"No news for {coin_name} between {week_start} and {week_end}.")
        else:
                logging.debug(f"Failed to fetch news for {coin_name} between {week_start} and {week_end}. Status Code: {response.status_code}")
            
        time.sleep(1)
        end_date -= timedelta(days=1)

    return pd.DataFrame(all_news)

def save_result_to_csv(result):
    """
    Saves a single result as a row in a CSV file.

    The result will be appended to the existing file if it exists, or written to a new file if not.

    Parameters:
        result (dict): A dictionary containing at least the keys 'coin', 'market_cap', 'volume_24h', 'price_change_7d', and 'fear_greed_index'.
    """
    if not os.path.exists(RESULTS_FILE):
        # If the file doesn't exist, create it with headers
        pd.DataFrame([result]).to_csv(RESULTS_FILE, mode='w', header=True, index=False)
    else:
        # If the file exists, append to it without writing headers again
        pd.DataFrame([result]).to_csv(RESULTS_FILE, mode='a', header=False, index=False)


def gpt4o_summarize_digest_and_extract_tickers(digest_text):
    """
    Uses GPT-4 to summarize the Sundown Digest and extract key points related to potential surges in coin value.

    Args:
        digest_text (str): The concatenated text from all digest entries.

    Returns:
        dict: A dictionary containing a summary focused on surge-causing news and a list of extracted tickers.
    """
    prompt = f"""
    Analyze the following digest entries and provide the following:
    1) A concise summary in bullet points (no more than 250 words) of key news items likely to cause surges in the value of the mentioned coins. 
    2) List the relevant cryptocurrency tickers beside each news item. Ensure there is no duplication.

    Text:
    {digest_text}

    Respond **only** in JSON format with 'surge_summary' and 'tickers' as keys. Ensure the tickers are in alphabetical order and there are no duplicate tickers.
    """

    try:
        response = openai.ChatCompletion.create(
            model="gpt-4o-mini",
            messages=[{"role": "user", "content": prompt}],
            max_tokens=400,
            n=1,
            stop=None,
            temperature=0.7
        )
        
        # Extract the content of the response
        response_content = response.choices[0].message['content'].strip()
        # Use a regular expression to extract JSON from the response content
        json_match = re.search(r'\{.*\}', response_content, re.DOTALL)
        if json_match:
            json_str = json_match.group(0)
            try:
                analysis = json.loads(json_str)
                return analysis
            except json.JSONDecodeError:
                logging.debug(f"Failed to decode JSON: {json_str}")
                return {"surge_summary": "", "tickers": []}
        else:
            logging.debug(f"No JSON found in the response: {response_content}")
            return {"surge_summary": "", "tickers": []}
        
    except openai.error.RateLimitError as e:
        logging.debug(f"Rate limit reached: {e}. Waiting for 60 seconds before retrying...")
        time.sleep(60)  # Wait before retrying
        return gpt4o_summarize_digest_and_extract_tickers(digest_text)  # Retry the request

    except Exception as e:
        logging.debug(f"An error occurred while summarizing the digest and extracting tickers: {e}")
        return {"surge_summary": "", "tickers": []}


def summarize_sundown_digest(digest):
    """
    Summarizes the Sundown Digest content from the last three days, including sentiment detection,
    coin ticker extraction, and a summary focused on news likely to cause surges in coin value.

    Args:
        digest (list): List of Sundown Digest entries.

    Returns:
        dict: A dictionary containing key points of news that may cause surges and relevant tickers.
    """
    # Get the current date and calculate the date three days ago
    current_date = datetime.now()
    three_days_ago = current_date - timedelta(days=3)

    digest_texts = []
    
    for entry in digest:
        # Parse the entry's date
        entry_date = datetime.strptime(entry['date'], '%Y-%m-%dT%H:%M:%S.%fZ')

        # Filter out entries older than three days
        if entry_date < three_days_ago:
            continue

        digest_texts.append(entry['text'])

    # Concatenate all digest texts into a single string
    combined_digest_text = " ".join(digest_texts)

    # Use GPT-4 to analyze and summarize the combined digest text
    summary_and_tickers = gpt4o_summarize_digest_and_extract_tickers(combined_digest_text)
    return summary_and_tickers


def get_sundown_digest():
    """
    Fetches the Sundown Digest from CryptoNews API.

    Returns:
        dict: A dictionary containing the Sundown Digest data.
    """
    url = f"https://cryptonews-api.com/api/v1/sundown-digest?page=1&token={os.getenv('CRYPTO_NEWS_API_KEY')}"
    try:
        response = requests.get(url)
        response.raise_for_status()
        sundown_digest = response.json()
        return sundown_digest.get('data', [])
    except requests.exceptions.HTTPError as http_err:
        logging.debug(f"HTTP error occurred: {http_err}")
    except Exception as err:
        logging.debug(f"An error occurred: {err}")
    return []


def fetch_santiment_metric(metric, coin_slug, start_date, end_date):
    """
    Fetches Santiment metric if it's part of the free metrics plan.
    If fetching fails, returns None, allowing the rest of the code to continue.

    Parameters:
        metric (str): The metric to fetch.
        coin_slug (str): The Santiment slug of the cryptocurrency.
        start_date (str): Start date in YYYY-MM-DD format.
        end_date (str): End date in YYYY-MM-DD format.

    Returns:
        float or None: The fetched data if available, or None if the API call fails.
    """
    try:
        logging.debug(f"Fetching Santiment metric: {metric} for coin: {coin_slug}")
        result = san.get(metric, slug=coin_slug, from_date=start_date, to_date=end_date)
        if not result.empty:
            # Extract the latest value
            return result.iloc[-1]['value']
        else:
            logging.debug(f"No data found for metric {metric} for coin {coin_slug}")
            return None
    except Exception as e:
        logging.error(f"Error fetching Santiment metric: {metric} for {coin_slug}: {e}")
        return None


def fetch_santiment_data_for_coin(coin_slug):
    """
    Fetches relevant Santiment data (social volume, dev activity, daily active addresses, etc.) for a specific coin.
    If the Santiment API fails, returns an empty result, allowing the rest of the code to execute.

    Parameters:
        coin_slug (str): The Santiment slug of the cryptocurrency.

    Returns:
        dict: A dictionary with fetched Santiment metrics data or default values if the API call fails.
    """
    try:
        end_date = datetime.now().strftime('%Y-%m-%d')
        start_date = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')

        # Fetch development activity
        dev_activity_increase = fetch_santiment_metric('30d_moving_avg_dev_activity_change_1d', coin_slug, start_date, end_date)
        dev_activity_increase = dev_activity_increase.iloc[-1]['value'] if not dev_activity_increase.empty else 0.0

        # Fetch daily active addresses
        daily_active_addresses_increase = fetch_santiment_metric('active_addresses_24h_change_30d', coin_slug, start_date, end_date)
        daily_active_addresses_increase = daily_active_addresses_increase.iloc[-1]['value'] if not daily_active_addresses_increase.empty else 0.0

        return {
            "dev_activity_increase": float(dev_activity_increase),
            "daily_active_addresses_increase": float(daily_active_addresses_increase),
        }

    except Exception as e:
        logging.error(f"Error fetching Santiment data for {coin_slug}: {e}")
        # Return default values if there's an error
        return {
            "dev_activity_increase": 0.0,
            "daily_active_addresses_increase": 0.0,
        }


def create_coin_data_table_if_not_exists():
    """
    Creates the 'coin_data' table in Amazon Aurora (PostgreSQL) if it doesn't already exist,
    storing time series data for cumulative scores.
    """
    connection = None  # Initialize the connection variable to None
    try:
        # Connect to PostgreSQL Aurora instance
        connection = psycopg2.connect(
            host=os.getenv('AURORA_HOST'),
            database=os.getenv('AURORA_DB'),
            user=os.getenv('AURORA_USER'),
            password=os.getenv('AURORA_PASSWORD'),
            port=os.getenv('AURORA_PORT', 5432)  # Default port for PostgreSQL is 5432
        )
        
        cursor = connection.cursor()

        # SQL to create the table if it doesn't exist, allowing time series data
        create_table_query = """
        CREATE TABLE IF NOT EXISTS coin_data (
            id SERIAL PRIMARY KEY,
            coin_id VARCHAR(255) NOT NULL,
            coin_name VARCHAR(255) NOT NULL,
            cumulative_score FLOAT NOT NULL,
            timestamp DATE DEFAULT CURRENT_DATE,
            UNIQUE (coin_id, timestamp)  -- Unique constraint to ensure one entry per coin per day
        );
        """
        cursor.execute(create_table_query)
        connection.commit()
        print("Table created or already exists.")

    except OperationalError as e:
        print(f"Error while connecting to Amazon Aurora: {e}")
    
    finally:
        # Close the connection if it was successfully created
        if connection:
            cursor.close()
            connection.close()
            print("PostgreSQL connection is closed.")

def monitor_coins_and_send_report():
    """
    Main entry point to monitor the specified coins, fetch news, analyze sentiment,
    and send a report with the results.

    If TEST_ONLY is set to True, only a few coins are processed and the results are
    saved to a file. Otherwise, all coins are processed and the results are sent
    via email.
    """
    # Ensure the table is created in Amazon Aurora if it doesn't exist
    create_coin_data_table_if_not_exists()

    if TEST_ONLY:
        existing_results = pd.DataFrame([])
        # Add a predefined mix of small, medium, and large-cap coins
        coins_to_monitor = [
            {"id": "btc-bitcoin", "name": "Bitcoin"},
            {"id": "eth-ethereum", "name": "Ethereum"},
        ]
    else:
        existing_results = load_existing_results()
        coins_to_monitor = api_call_with_retries(client.coins)
        logging.debug(f"Number of coins retrieved: {len(coins_to_monitor)}")
        coins_to_monitor = filter_active_and_ranked_coins(coins_to_monitor, 1000)

    logging.debug(f"Number of active and ranked coins selected: {len(coins_to_monitor)}")
    end_date = datetime.now().strftime('%Y-%m-%d')

    report_entries = []
    tickers_dict = load_tickers(CRYPTO_NEWS_TICKERS)

    # Fetch and summarize the Sundown Digest
    sundown_digest = get_sundown_digest()
    digest_summary = summarize_sundown_digest(sundown_digest)
    digest_tickers = digest_summary['tickers']

    # Fetch Trending Coins data once
    trending_coins_scores = fetch_trending_coins_scores()

    # Load Santiment slugs
    santiment_slugs_df = fetch_santiment_slugs()

    for coin in coins_to_monitor:
        try:
            print(f"Processing {coin['name']} ({coin['id']})")
            coin_id = coin['id']
            coin_name = coin['name'].lower()

            if not existing_results.empty and coin_id in existing_results['coin_id'].values:
                logging.debug(f"Skipping already processed coin: {coin_id}")
                continue

            # Fetch news directly for analysis
            coins_dict = {coin_name: tickers_dict.get(coin_name, '').upper()}
            news_df = fetch_news_for_past_week(coins_dict)

            # Analyze coin and save the result, passing the Santiment slug
            result = analyze_coin(coin_id, coin_name, end_date, news_df, digest_tickers, trending_coins_scores, santiment_slugs_df)
            logging.debug(f"Result for {coin_name}: {result}")

            save_result_to_csv(result)
            report_entries.append(result)

            # Save the cumulative score to Amazon Aurora
            save_cumulative_score_to_aurora(result['coin_id'], result['coin_name'], result['cumulative_score_percentage'])

            time.sleep(20)

        except Exception as e:
            logging.debug(f"An error occurred while processing {coin_name} ({coin_id}): {e}")
            logging.debug(traceback.format_exc())
            continue

    df = pd.DataFrame(report_entries)

    if not df.empty:
        # Filter the report_entries based on liquidity risk and cumulative score percentage
        df = df[(df['liquidity_risk'].isin(['Low', 'Medium'])) & (df['cumulative_score_percentage'] > CUMULATIVE_SCORE_REPORTING_THRESHOLD)]

        logging.debug("DataFrame is not empty, processing report entries.")

        # Extract the coin names from the filtered DataFrame
        coins_in_df = df['coin_name'].unique()  # Extract unique coin names from the filtered DataFrame

        if len(coins_in_df) > 0:
            # Retrieve historical data from Amazon Aurora
            historical_data = retrieve_historical_data_from_aurora()

            if not historical_data.empty:
                # Filter the historical data for only the coins present in the filtered df
                plot_top_coins_over_time(historical_data[historical_data['coin_name'].isin(coins_in_df)], top_n=len(coins_in_df))

        # Proceed with sorting and generating the report
        report_entries = df.to_dict('records')
        report_entries = sorted(report_entries, key=lambda x: x.get('cumulative_score', 0), reverse=True)
        logging.debug(f"Report entries after sorting: {report_entries}")

        logging.debug(f"DataFrame contents before GPT-4o recommendations:\n{df.to_string()}")

        # Get GPT-4o recommendations
        gpt_recommendations = gpt4o_analyze_and_recommend(df)
        logging.debug(f"GPT-4o recommendations: {gpt_recommendations}")

        # Generate HTML report with recommendations
        html_report = generate_html_report_with_recommendations(report_entries, digest_summary, gpt_recommendations)
        logging.debug("HTML report generated successfully.")

        # Save the report to Excel
        attachment_path = save_report_to_excel(report_entries)
        logging.debug(f"Report saved to Excel at: {attachment_path}")

        # Send email with the report and the plot attached
        send_email_with_report(html_report, attachment_path)
        logging.debug("Email sent successfully.")

        # Delete the surging_coins.csv file after sending the email
        if os.path.exists(RESULTS_FILE):
            try:
                os.remove(RESULTS_FILE)
                logging.debug(f"{RESULTS_FILE} has been deleted successfully.")
            except Exception as e:
                logging.debug(f"Failed to delete {RESULTS_FILE}: {e}")
    else:
        logging.debug("No valid entries to report. DataFrame is empty.")

def normalize_score(raw_score, min_score, max_score, range=3):
    """
    Normalize a score to be between 0 and 3 and return the nearest integer.

    Args:
        raw_score (float): The original score.
        min_score (float): The minimum score in the data.
        max_score (float): The maximum score in the data.

    Returns:
        int: The normalized score between 0 and 3, rounded to the nearest integer.
    """
    # Handle the edge case where all scores might be the same
    if max_score == min_score:
        return 2  # Midpoint value since all scores are the same

    # Normalize to a 0-1 range
    normalized = (raw_score - min_score) / (max_score - min_score)
    # Scale to a 0-3 range and round to the nearest integer
    return round(normalized * range)


def fetch_trending_coins_scores():
    """
    Fetches trending coins from CryptoNews API and calculates their sentiment scores.
    Normalizes the scores between 0 and 3 and returns a dictionary with the trending coins and their normalized scores.

    Returns:
        dict: A dictionary with the trending coins as keys and their normalized scores as values.
    """
    url = f"https://cryptonews-api.com/api/v1/top-mention?&date=last7days&token={os.getenv('CRYPTO_NEWS_API_KEY')}"
    response = requests.get(url)
    trending_data = response.json()['data']['all']

    trending_coins_scores = {}
    raw_scores = {}

    # Calculate raw scores
    for item in trending_data:
        ticker = item['ticker'].lower()
        sentiment_score = item['sentiment_score']
        total_mentions = item['total_mentions']
        raw_score = sentiment_score * total_mentions
        raw_scores[ticker] = raw_score

    # Determine min and max raw scores for normalization
    min_raw_score = min(raw_scores.values())
    max_raw_score = max(raw_scores.values())

    # Normalize scores between 0 and 3
    for ticker, raw_score in raw_scores.items():
        trending_coins_scores[ticker] = normalize_score(raw_score, min_raw_score, max_raw_score)

    return trending_coins_scores


def send_failure_email():
    """
    Sends an email with the current results when the script encounters an error.

    This function reads the current results from the file specified by RESULTS_FILE,
    generates an HTML email with the results and sends it to the recipient specified
    by EMAIL_TO.

    If the file does not exist, the email will state that no data is available.

    The email is sent using the SMTP server specified by SMTP_SERVER and the
    credentials specified by SMTP_USERNAME and SMTP_PASSWORD.

    If sending the email fails, an error message is logging.debuged.
    """
    if os.path.exists(RESULTS_FILE):
        with open(RESULTS_FILE, 'r') as file:
            file_contents = file.read()
    else:
        file_contents = "No data available, as the results file was not created."

    # HTML content with inline CSS for the failure email
    html_content = f"""
    <html>
    <head>
        <style>
            body {{
                font-family: Arial, sans-serif;
                color: #333;
            }}
            h2 {{
                color: #c0392b;
            }}
            p {{
                font-size: 14px;
                color: #555;
            }}
            .content {{
                background-color: #f9f9f9;
                padding: 20px;
                border: 1px solid #ddd;
                border-radius: 5px;
            }}
            .content pre {{
                background-color: #f4f4f4;
                border: 1px solid #ccc;
                padding: 10px;
                border-radius: 3px;
            }}
        </style>
    </head>
    <body>
        <h2>Failure in Weekly Coin Analysis Script</h2>
        <p>The script encountered an error. Below are the current results:</p>
        <div class="content">
            <pre>{file_contents}</pre>
        </div>
    </body>
    </html>
    """

    msg = MIMEMultipart('alternative')
    msg['Subject'] = "Failure in Weekly Coin Analysis Script"
    msg['From'] = EMAIL_FROM
    msg['To'] = EMAIL_TO

    part = MIMEText(html_content, 'html')
    msg.attach(part)

    try:
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls()
            server.login(SMTP_USERNAME, SMTP_PASSWORD)
            server.sendmail(EMAIL_FROM, EMAIL_TO, msg.as_string())
        logging.debug("Failure email sent successfully.")
    except Exception as e:
        logging.debug(f"Failed to send email: {e}")


def get_price_change_thresholds(market_cap_class, volatility_class):
    """
    Returns the price change thresholds for a given market capitalization class and volatility class.

    Parameters:
        market_cap_class (str): The market capitalization class, one of "Large", "Mid", or "Small".
        volatility_class (str): The volatility class, one of "High", "Medium", or "Low".

    Returns:
        tuple: A tuple of three floats, representing the price change thresholds for short-term, medium-term, and long-term periods, respectively.
    """
    thresholds = {
        ("Large", "High"): (0.03, 0.02, 0.01),
        ("Large", "Low"): (0.015, 0.01, 0.005),
        ("Mid", "High"): (0.05, 0.03, 0.02),
        ("Mid", "Medium"): (0.03, 0.02, 0.015),
        ("Mid", "Low"): (0.02, 0.015, 0.01),
        ("Small", "High"): (0.07, 0.05, 0.03),
        ("Small", "Medium"): (0.05, 0.03, 0.02),
        ("Small", "Low"): (0.03, 0.02, 0.015)
    }
    return thresholds.get((market_cap_class, volatility_class), (0.03, 0.02, 0.01))

def analyze_price_change(price_data, market_cap, volatility):
    """
    Analyze the price changes of a cryptocurrency over three time periods.

    Parameters:
        price_data (pd.DataFrame): Price data for the cryptocurrency.
        market_cap (int): Market capitalization of the cryptocurrency.
        volatility (float): Volatility of the cryptocurrency.

    Returns:
        tuple: The price change score and an explanation string detailing
               which periods had significant price changes.
    """
    market_cap_class = classify_market_cap(market_cap)
    volatility_class = classify_volatility(volatility)

    # Unpack only the lower threshold values
    short_term_threshold, medium_term_threshold, long_term_threshold = get_price_change_thresholds(market_cap_class, volatility_class)

    # Analyze short-term, medium-term, and long-term price changes
    short_term_change = calculate_price_change(price_data, period="short")
    medium_term_change = calculate_price_change(price_data, period="medium")
    long_term_change = calculate_price_change(price_data, period="long")

    price_change_score = 0
    explanation_parts = []

    if short_term_change > short_term_threshold:
        price_change_score += 1
        explanation_parts.append(f"Short-term change of {short_term_change*100:.2f}% exceeded the threshold of {short_term_threshold*100:.2f}%")
    
    if medium_term_change > medium_term_threshold:
        price_change_score += 1
        explanation_parts.append(f"Medium-term change of {medium_term_change*100:.2f}% exceeded the threshold of {medium_term_threshold*100:.2f}%")
    
    if long_term_change > long_term_threshold:
        price_change_score += 1
        explanation_parts.append(f"Long-term change of {long_term_change*100:.2f}% exceeded the threshold of {long_term_threshold*100:.2f}%")

    explanation = " | ".join(explanation_parts) if explanation_parts else "No significant price changes detected."

    return price_change_score, explanation


def calculate_price_change(price_data, period="short", span=7):
    """
    Calculate the percentage change in price over a given period with optional smoothing using Exponential Moving Average (EMA).

    Parameters:
        price_data (pd.DataFrame): Price data for the cryptocurrency with a single column.
        period (str, optional): The time period to calculate the price change for. Options are "short" (7 days), "medium" (30 days), and "long" (90 days). Defaults to "short".
        span (int, optional): The span for the EMA smoothing. Defaults to 7.

    Returns:
        float: The percentage change in price over the specified period, smoothed by EMA.
    """
    # Apply EMA
    smoothed_data = price_data.ewm(span=span, adjust=False).mean()

    if period == "short":
        period_data = smoothed_data.tail(7)  # Last 7 days
    elif period == "medium":
        period_data = smoothed_data.tail(30)  # Last 30 days
    else:  # long term
        period_data = smoothed_data.tail(90)  # Last 90 days

    start_price = period_data.iloc[0]  # Extract the first value from the single column
    end_price = period_data.iloc[-1]  # Extract the last value from the single column

    if start_price == 0:
        return None  # or some other appropriate value or handling mechanism
    return (end_price - start_price) / start_price


def calculate_volume_change(volume_data, period="short", span=7):
    """
    Calculate the percentage change in volume over a given period with optional smoothing using Exponential Moving Average (EMA).

    Parameters:
        volume_data (pd.DataFrame): Volume data for the cryptocurrency with a single column.
        period (str, optional): The time period to calculate the volume change for. Options are "short" (7 days), "medium" (30 days), and "long" (90 days). Defaults to "short".
        span (int, optional): The span for the EMA smoothing. Defaults to 7.

    Returns:
        float: The percentage change in volume over the specified period, smoothed by EMA.
    """
    # Apply EMA
    smoothed_data = volume_data.ewm(span=span, adjust=False).mean()

    if period == "short":
        period_data = smoothed_data.tail(7)  # Last 7 days
    elif period == "medium":
        period_data = smoothed_data.tail(30)  # Last 30 days
    else:  # long term
        period_data = smoothed_data.tail(90)  # Last 90 days

    start_volume = period_data.iloc[0]  # Extract the first value from the single column
    end_volume = period_data.iloc[-1]  # Extract the last value from the single column

    if start_volume == 0:
        return None  # or some other appropriate value or handling mechanism
    return (end_volume - start_volume) / start_volume

def print_command_line_report(report_entries):
    """
    Prints a command-line report of the daily coin analysis.

    Parameters
    ----------
    report_entries : list
        A list of dictionaries, each containing the analysis results for a single coin.

    Returns
    -------
    None
    """
    df = pd.DataFrame(report_entries)
    logging.debug("\nCoin Analysis Report")
    logging.debug(tabulate(df, headers="keys", tablefmt="grid"))
    logging.debug(f"\nReport generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

def print_command_line_report(report_entries, digest_summary):
    """
    Prints a command-line report of the daily coin analysis, with the digest summary
    printed separately before the table.

    Parameters:
    report_entries (list): A list of dictionaries, each containing the analysis results for a single coin.
    digest_summary (dict): A dictionary containing the summarized Sundown Digest.
    
    Returns:
    None
    """
    # logging.debug the Sundown Digest summary first
    logging.debug("\nSundown Digest Summary:")

    logging.debug("\nTickers Mentioned:")
    logging.debug(", ".join(digest_summary['tickers']))

    logging.debug("\nNews Sumnmary:")
    for surge in digest_summary['surge_summary']:
        logging.debug(f"- {surge}")

    # Then logging.debug the table of analyzed coins
    df = pd.DataFrame(report_entries)
    logging.debug("\nCoin Analysis Report")
    logging.debug(tabulate(df, headers="keys", tablefmt="grid"))
    logging.debug(f"\nReport generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

def gpt4o_analyze_and_recommend(df):
    """
    Uses GPT-4o to analyze the final results DataFrame and provide structured recommendations for coin purchases.

    Parameters:
        df (pd.DataFrame): The final DataFrame containing coin analysis results.

    Returns:
        dict: A structured summary of GPT-4o's recommendations for coin purchases, including reasons.
    """
    # Convert the entire DataFrame to a JSON format for GPT-4o input
    df_json = df.to_dict(orient='records')

    # Adjust the prompt to request structured JSON output
    prompt = f"""
    You are provided with detailed analysis data for several cryptocurrency coins. Using this data, provide a concise summary of which coins should be considered for purchase, along with the reasons for the recommendation. Only consider coins that have a potential of a breakout or a surge in value.

    **Do not repeat or summarize the dataset.** Instead, return the recommendations in structured JSON format with each recommended coin and give a detailed reason for your recommendation based on the market data you have been given.

    Format your response as follows:
    {{
        "recommendations": [
            {{
                "coin": "Coin Name",
                "liquidity_risk": "Low/Medium/High",
                "cumulative_score": "Score Value",
                "reason": "A fluent and detailed reason for recommendation anchored in the data you have been provided."
            }},
            ...
        ]
    }}

    Here is the data for your analysis:
    {json.dumps(df_json, indent=2)}
    """

    try:
        # Call the OpenAI API to generate recommendations
        response = openai.ChatCompletion.create(
            model="gpt-4o-mini",
            messages=[{"role": "user", "content": prompt}],
            n=1,
            max_tokens=4000,
            stop=None,
            temperature=0.1
        )

        # Extract the response content (the actual GPT output text)
        gpt_message_content = response['choices'][0]['message']['content']

        # Log the GPT response for debugging
        logging.debug("Response_content: " + gpt_message_content)
        
        # Extract the JSON part using regex
        json_match = re.search(r'```json(.*?)```', gpt_message_content, re.DOTALL)

        # Check if JSON part is found and not empty
        if json_match:
            json_content = json_match.group(1).strip()

            # Check if the extracted content is empty
            if not json_content:
                logging.debug("Error: JSON content is empty.")
            else:
                # Try to parse the JSON content
                try:
                    parsed_data = json.loads(json_content)
                    logging.debug("Parsed JSON data:", parsed_data)
                except json.JSONDecodeError as e:
                    logging.debug(f"Failed to parse JSON: {e}")
        else:
            logging.debug("No JSON found in the log message.")
     
        # Check if the 'recommendations' field is empty
        if not parsed_data['recommendations']:
            logging.debug("No recommendations found in the response.")
        else:
            logging.debug(f"Recommendations found: {parsed_data['recommendations']}")

        return parsed_data

    except (openai.error.OpenAIError, json.JSONDecodeError) as e:
        logging.debug(f"Error in GPT-4o analysis: {e}")
        return {"recommendations": []}


def generate_html_report_with_recommendations(report_entries, digest_summary, gpt_recommendations, plot_image_path='top_coins_plot.png'):
    """
    Generates an HTML report with summaries from the report entries, GPT-4o recommendations, and a plot of the top coins.

    Args:
        report_entries (list): List of report entries to include in the report.
        digest_summary (dict): Summary of the Sundown Digest to include at the top.
        gpt_recommendations (dict): GPT-4o's recommendations for coin purchases, structured as a list of dictionaries.
        plot_image_path (str): Path to the plot image to embed in the report.

    Returns:
        str: HTML content of the report.
    """
    
    # Sundown Digest Summary section
    digest_items = ''.join(f'<li style="font-size:14px;line-height:1.6;">{item}</li>' for item in digest_summary['surge_summary'])
    digest_html = f"""
    <table width="100%" cellpadding="0" cellspacing="0" border="0" style="background-color:#fff;">
        <tr>
            <td style="padding:20px;">
                <h3 style="font-size:20px;color:#2a9d8f;margin-bottom:10px;">Sundown Digest Summary</h3>
                <p style="font-size:14px;line-height:1.6;"><strong>Tickers Mentioned:</strong> {', '.join(digest_summary['tickers'])}</p>
                <p style="font-size:14px;line-height:1.6;"><strong>News Summary:</strong></p>
                <ul style="list-style-type:disc;padding-left:20px;margin:0;">
                    {digest_items}
                </ul>
            </td>
        </tr>
    </table>
    """

    # AI Recommendations Section
    if not gpt_recommendations['recommendations']:
        recommendations_html = """
        <table width="100%" cellpadding="0" cellspacing="0" border="0" style="background-color:#fff;">
            <tr>
                <td style="padding:20px;">
                    <h3 style="font-size:20px;color:#2a9d8f;margin-bottom:10px;">AI Generated Coin Recommendations</h3>
                    <p style="font-size:14px;line-height:1.6;">No coins are currently recommended for purchase based on the analysis.</p>
                </td>
            </tr>
        </table>
        """
    else:
        recommendation_items = ''
        for item in gpt_recommendations['recommendations']:
            # Match the coin with report entries to fetch URL, cumulative score percentage
            matching_entry = next((entry for entry in report_entries if entry["coin_name"].lower() == item["coin"].lower()), None)
            
            # CoinPaprika URL format or other URL source can be used here
            coin_url = f"https://coinpaprika.com/coin/{matching_entry['coin_id']}/" if matching_entry else '#'
            cumulative_score_percentage = matching_entry.get('cumulative_score_percentage', 'N/A')

            # Capitalize each word in the coin name
            coin_name = item["coin"].title()

            recommendation_items += f"""
            <li style="font-size:14px;line-height:1.6;margin-bottom:10px;">
                <b>{coin_name}</b> - {item["reason"]}<br>
                <strong>Cumulative Score Percentage:</strong> {cumulative_score_percentage}%<br>
                <a href="{coin_url}" target="_blank" style="color:#0077cc;text-decoration:none;">More Info</a>
            </li>
            """
        recommendations_html = f"""
        <table width="100%" cellpadding="0" cellspacing="0" border="0" style="background-color:#fff;">
            <tr>
                <td style="padding:20px;">
                    <h3 style="font-size:20px;color:#2a9d8f;margin-bottom:10px;">AI Generated Coin Recommendations</h3>
                    <p style="font-size:14px;line-height:1.6;"><strong>Meaning of Cumulative Score Percentage:</strong> a higher percentage indicates a stronger potential based on historical data and analysis. </p>
                    <ul style="list-style-type:disc;padding-left:20px;margin:0;">
                        {recommendation_items}
                    </ul>
                </td>
            </tr>
        </table>
        """

    # Embed the attached image in the HTML using CID
    cid = "top_coins_plot"  # This should match the Content-ID of the attached image
    plot_html = f"""
    <table width="100%" cellpadding="0" cellspacing="0" border="0" style="background-color:#fff;">
        <tr>
            <td style="padding:20px;text-align:center;">
                <h3 style="font-size:20px;color:#2a9d8f;margin-bottom:10px;">Top Coins Cumulative Scores Over Time</h3>
                <img src="cid:{cid}" alt="Top Coins Plot" style="width:100%;max-width:600px;height:auto;"/>
            </td>
        </tr>
    </table>
    """

    # Full HTML structure
    html_content = f"""
    <html>
    <body style="margin:0;padding:0;background-color:#f9f9f9;font-family:Arial,sans-serif;color:#333;">
        <table width="100%" cellpadding="0" cellspacing="0" border="0" style="background-color:#f9f9f9;">
            <tr>
                <td align="center">
                    <table width="600" cellpadding="0" cellspacing="0" border="0" style="background-color:#fff;">
                        <tr>
                            <td style="padding:20px;">
                                <h2 style="text-align:center;color:#264653;font-size:24px;margin:0;">Coin Analysis Report</h2>
                            </td>
                        </tr>
                        <tr>
                            <td>
                                {digest_html}
                            </td>
                        </tr>
                        <tr>
                            <td>
                                {recommendations_html}
                            </td>
                        </tr>
                        <tr>
                            <td>
                                {plot_html}
                            </td>
                        </tr>
                        <tr>
                            <td style="padding:20px;">
                                <p style="text-align:center;color:#777;font-size:12px;margin:0;">Report generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
                            </td>
                        </tr>
                    </table>
                </td>
            </tr>
        </table>
    </body>
    </html>
    """

    return html_content


def save_report_to_excel(report_entries, filename='coin_analysis_report.xlsx'):
    """
    Saves the report entries to an Excel file with enhanced formatting and styling.

    Args:
        report_entries (list): A list of dictionaries containing the report data.
        filename (str): The name of the Excel file to save the report to.
    """
    # Convert the report entries to a pandas DataFrame
    df = pd.DataFrame(report_entries)
    
    # Save DataFrame to an Excel file without formatting
    df.to_excel(filename, index=False)
    
    # Open the Excel file with openpyxl for formatting
    workbook = load_workbook(filename)
    sheet = workbook.active

    # Define styles for headers and cells
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="4F81BD")
    cell_font = Font(name="Arial", size=10)
    cell_alignment = Alignment(horizontal="left", vertical="top", wrap_text=False)  # Turn off wrap_text for content cells
    
    # Define border style
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                         top=Side(style="thin"), bottom=Side(style="thin"))

    # Apply header styles (background color, font, alignment)
    for col in sheet.iter_cols(min_row=1, max_row=1, min_col=1, max_col=sheet.max_column):
        max_length = 0
        column = col[0].column_letter  # Get the column letter for header
        for cell in col:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)  # Turn wrapping off for headers
            cell.border = thin_border
            # Adjust column width based on header content
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))

        adjusted_width = (max_length + 2) * 1.2  # Add some padding for headers
        sheet.column_dimensions[column].width = adjusted_width

    # Apply cell styles (font, alignment, borders) and auto-adjust column width based on content
    for col in sheet.iter_cols(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        max_length = 0
        column = col[0].column_letter  # Get the column letter for data cells

        for cell in col:
            cell.font = cell_font
            cell.alignment = cell_alignment
            cell.border = thin_border

            # Adjust column width based on the content
            try:
                if cell.value:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
            except Exception as e:
                print(f"Error processing cell {cell.coordinate}: {e}")

        # Set the column width to fit the content with padding
        adjusted_width = (max_length + 2) * 1.2  # Add padding for cells
        sheet.column_dimensions[column].width = adjusted_width

    # Freeze the top row (headers) for better readability
    sheet.freeze_panes = "A2"

    # Save the workbook with the formatting applied
    try:
        workbook.save(filename)
        print(f"Report saved to {filename} with enhanced formatting.")
    except Exception as e:
        print(f"Error saving the report: {e}")
    finally:
        workbook.close()

    return filename


def plot_top_coins_over_time(historical_data, top_n=5, file_name='top_coins_plot.png'):
    """
    Plots the cumulative scores of the top coins over time and saves the plot to a file.

    Args:
        historical_data (pd.DataFrame): DataFrame containing the historical data with 'coin_name', 'cumulative_score', and 'timestamp' columns.
        top_n (int): The number of top coins to plot.
        file_name (str): The name of the file to save the plot to.
    """
    # Avoid the SettingWithCopyWarning by using .loc
    historical_data.loc[:, 'timestamp'] = pd.to_datetime(historical_data['timestamp'])

    # Calculate the average cumulative score for each coin and select the top N coins
    top_coins = historical_data.groupby('coin_name')['cumulative_score'].mean().nlargest(top_n).index

    # Filter data for only the top coins
    top_data = historical_data[historical_data['coin_name'].isin(top_coins)]

    # Plot each top coin's cumulative score over time
    plt.figure(figsize=(10, 6))
    for coin in top_coins:
        coin_data = top_data[top_data['coin_name'] == coin]
        # Add markers to the plot (e.g., 'o' for circles)
        plt.plot(coin_data['timestamp'], coin_data['cumulative_score'], label=coin, marker='o')

    # Format x-axis with date formatting based on the range of dates in the data
    plt.gca().xaxis.set_major_locator(mdates.AutoDateLocator())  # Automatically adjust the date ticks
    plt.gca().xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m-%d'))  # Format the ticks as 'Year-Month-Day'
    
    # Plot settings
    plt.title(f'Top {top_n} Coins by Cumulative Score Over Time')
    plt.xlabel('Date')
    plt.ylabel('Cumulative Score')
    plt.legend()

    # Save plot to file
    plt.tight_layout()
    plt.savefig(file_name)
    #plt.show()


def retrieve_historical_data_from_aurora():
    """
    Retrieves historical cumulative scores from Amazon Aurora for all coins.

    Returns:
        pd.DataFrame: A DataFrame containing the timestamp, coin name, and cumulative score.
    """
    engine = None
    try:
        # Build the database connection string
        db_connection_str = (
            f"postgresql://{os.getenv('AURORA_USER')}:{os.getenv('AURORA_PASSWORD')}"
            f"@{os.getenv('AURORA_HOST')}:{os.getenv('AURORA_PORT', 5432)}/{os.getenv('AURORA_DB')}"
        )

        # Create an SQLAlchemy engine
        engine = create_engine(db_connection_str)

        # Define the SQL query to retrieve time series data
        query = """
            SELECT coin_name, cumulative_score, timestamp 
            FROM coin_data
            ORDER BY timestamp;
        """
        
        # Use pandas to execute the query and return the result as a DataFrame
        df = pd.read_sql(query, engine)
        print("Historical data retrieved successfully.")
        return df

    except SQLAlchemyError as e:
        print(f"Error retrieving historical data: {e}")
        return pd.DataFrame()  # Return empty DataFrame on failure

    finally:
        if engine:
            engine.dispose()  # Close the connection
            print("PostgreSQL connection is closed.")

def send_email_with_report(html_content, attachment_path, plot_image_path='top_coins_plot.png'):
    """
    Sends an email with an HTML report and an attached image.

    The email uses a 'related' MIME type to allow both HTML and images to be attached.
    The HTML content is passed as a string and the image is attached as an inline
    attachment with a Content-ID header that matches the CID in the HTML content.

    Args:
        html_content (str): The HTML content of the email.
        attachment_path (str): The path to the image file to attach.

    Returns:
        None
    """
    msg = MIMEMultipart('related')  # 'related' allows attaching both HTML and images
    msg['Subject'] = "AI Generated Coin Analysis Report"
    msg['From'] = EMAIL_FROM
    msg['To'] = EMAIL_TO

    # Attach HTML content
    part = MIMEText(html_content, 'html')
    msg.attach(part)

    # Attach the image (inline attachment with Content-ID)
    with open(plot_image_path, 'rb') as img_file:
        mime_image = MIMEImage(img_file.read(), _subtype='.png')
        mime_image.add_header('Content-ID', '<top_coins_plot>')  # Content-ID should match CID in HTML
         # Set Content-Disposition with a filename
        mime_image.add_header('Content-Disposition', 'inline', filename="top_coins_plot.png")
        msg.attach(mime_image)

    # Send the email
    with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
        server.starttls()
        server.login(SMTP_USERNAME, SMTP_PASSWORD)
        server.sendmail(EMAIL_FROM, EMAIL_TO, msg.as_string())


if __name__ == "__main__":
    monitor_coins_and_send_report()
